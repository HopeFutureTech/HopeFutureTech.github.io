#!/usr/bin/env python3
"""将 Excel 课程数据转换为网站所需的 JSON 格式。"""

import argparse
import json
import os
from collections import defaultdict

import pandas as pd

EXCEL_CANDIDATES = [
    os.path.join(os.path.dirname(__file__), '高一秋季全科系统课场次数据.xlsx'),
    os.path.join(os.path.dirname(__file__), '高一秋季全科系统课场次数据new.xlsx'),
]
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'data', 'courses.json')
SUBJECT_ORDER = ['语文', '数学', '英语', '物理', '化学', '生物', '历史', '地理', '政治']
YELLOW_RGB = 'FFFFFF00'
# 新版 Excel 列名与旧版映射
COLUMN_ALIASES = {
    '学科名称': ('学科名称', '学科'),
    '学期名称': ('学期名称', '学期'),
    '学年名称': ('学年名称', '学年'),
    '课程ID': ('课程ID',),
    '课程名称': ('课程名称',),
    '主讲老师名称': ('主讲老师名称', '师名'),
    '场次ID': ('场次ID',),
    '场次名称': ('场次名称',),
    '课程视频': ('课程视频', '视频'),
}


def get_excel_path():
    existing = [p for p in EXCEL_CANDIDATES if os.path.isfile(p)]
    if not existing:
        raise FileNotFoundError('未找到 Excel 文件，请将数据文件放在 highschool 目录下')
    return max(existing, key=os.path.getmtime)


def normalize_columns(df):
    """统一新旧 Excel 列名。"""
    rename = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in df.columns:
                rename[alias] = canonical
                break
    missing = [k for k in COLUMN_ALIASES if k not in rename.values() and k not in df.columns]
    if missing:
        raise ValueError(f'Excel 缺少必要列: {missing}，当前列: {list(df.columns)}')
    return df.rename(columns=rename)


def remove_yellow_rows(excel_path=None):
    """删除 Excel 中背景标黄（FFFFFF00）的数据行。"""
    from openpyxl import load_workbook

    excel_path = excel_path or get_excel_path()

    def is_yellow(cell):
        fill = cell.fill
        if fill is None or fill.fill_type != 'solid':
            return False
        fg = fill.fgColor
        if fg and fg.type == 'rgb' and fg.rgb:
            return str(fg.rgb).upper().replace(' ', '') == YELLOW_RGB
        return False

    wb = load_workbook(excel_path)
    ws = wb.active
    yellow_rows = [
        r for r in range(2, ws.max_row + 1)
        if any(is_yellow(ws.cell(r, c)) for c in range(1, ws.max_column + 1))
    ]
    for r in sorted(yellow_rows, reverse=True):
        ws.delete_rows(r)
    wb.save(excel_path)
    print(f'已从 {excel_path} 删除 {len(yellow_rows)} 行标黄数据')
    return len(yellow_rows)


def convert():
    excel_path = get_excel_path()
    print(f'读取: {excel_path}')
    df = normalize_columns(pd.read_excel(excel_path))
    data = defaultdict(lambda: defaultdict(lambda: {
        'courseId': None,
        'courseName': '',
        'teacher': '',
        'term': '',
        'year': '',
        'sessions': []
    }))
    seen_sessions = set()

    for _, row in df.iterrows():
        subj = row['学科名称']
        course_id = str(row['课程ID'])
        session_key = (subj, course_id, row['场次ID'])

        if session_key in seen_sessions:
            continue
        seen_sessions.add(session_key)

        if data[subj][course_id]['courseId'] is None:
            data[subj][course_id].update({
                'courseId': int(row['课程ID']),
                'courseName': row['课程名称'],
                'teacher': row['主讲老师名称'],
                'term': row['学期名称'],
                'year': row['学年名称'],
            })

        data[subj][course_id]['sessions'].append({
            'id': int(row['场次ID']),
            'name': row['场次名称'],
            'video': row['课程视频'],
        })

    result = {}
    for subj in SUBJECT_ORDER:
        if subj not in data:
            continue
        courses = sorted(data[subj].values(), key=lambda x: x['courseName'])
        result[subj] = courses

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False)

    total = sum(len(c['sessions']) for courses in result.values() for c in courses)
    print(f'已生成 {OUTPUT_PATH}')
    print(f'科目: {len(result)}, 课程: {sum(len(c) for c in result.values())}, 场次: {total}')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Excel 转 JSON，可选删除标黄行')
    parser.add_argument('--remove-yellow', action='store_true', help='删除 Excel 中背景标黄的行后再转换')
    args = parser.parse_args()
    if args.remove_yellow:
        remove_yellow_rows()
    convert()
